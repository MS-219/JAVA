#!/usr/bin/env python3
"""
把 collected_records.json 转换为 Excel，并通过网关 API 写入 H2 数据库（大屏恢复）
"""
import json
import os
import urllib.request

# 设备 MAC -> 名称映射
DEVICE_MAP = {
    "61:B2": "大门东-中间 出口",
    "61:CD": "大门东-东侧 入口",
    "61:A5": "大门东-西侧 入口",
    "61:D2": "大门西-东侧 入口",
    "61:CF": "大门西-西侧 出口",
    "61:D5": "大门西-中间 入口",
}

def get_device_name(mac):
    if not mac:
        return "未知设备"
    mac_upper = mac.upper()
    for suffix, name in DEVICE_MAP.items():
        if suffix in mac_upper:
            return name
    return f"设备 {mac[-5:]}" if len(mac) > 5 else f"设备 {mac}"

def get_direction(device_name):
    if "入口" in device_name:
        return "入口"
    elif "出口" in device_name:
        return "出口"
    return ""

def get_location(device_name):
    if device_name.endswith(" 入口") or device_name.endswith(" 出口"):
        return device_name.rsplit(" ", 1)[0]
    return device_name

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    input_file = os.path.join(base_dir, "collected_records.json")
    
    # 读取记录
    records = []
    seen = set()  # 去重
    with open(input_file, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                r = json.loads(line)
                # 去重 key: mac + employee_number + time
                key = (r.get("Mac_addr", ""), r.get("employee_number", ""), r.get("time", ""))
                if key in seen:
                    continue
                seen.add(key)
                records.append(r)
            except:
                continue
    
    print(f"读取到 {len(records)} 条记录 (去重后)")
    
    # 按时间排序
    records.sort(key=lambda x: x.get("time", ""), reverse=True)
    
    # ========== 生成 Excel ==========
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "识别记录 4月17日-23日"

    headers = ["序号", "姓名", "人员编号", "识别时间", "设备位置", "进出方向", "识别方式", "比对结果", "MAC地址"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    id_type_map = {"0": "人脸识别", "1": "黑名单", "2": "人证比对", "3": "IC卡"}
    result_map = {1: "成功", 0: "失败"}

    for idx, r in enumerate(records, 1):
        mac = r.get("Mac_addr", "")
        device_name = get_device_name(mac)
        
        row_data = [
            idx,
            r.get("name", "未知"),
            r.get("employee_number", ""),
            r.get("time", ""),
            get_location(device_name),
            get_direction(device_name),
            id_type_map.get(str(r.get("IdentifyType", "0")), "人脸识别"),
            result_map.get(r.get("resultStatus", 1), "成功"),
            mac,
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 1, column=col, value=val)
            cell.border = thin_border
            if col in [1, 5, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="center")

    # 列宽
    widths = [8, 12, 18, 22, 20, 10, 12, 10, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    output_file = os.path.join(base_dir, "识别记录_2026-04-17_至_2026-04-23.xlsx")
    wb.save(output_file)
    print(f"\n✅ Excel 已保存: {output_file}")
    print(f"   共 {len(records)} 条记录")

    # ========== 写入 H2 数据库 (通过网关 API) ==========
    print("\n开始写入大屏数据库...")
    gateway_url = "http://10.10.245.252:5389/device/record"
    success = 0
    fail = 0
    
    for i, r in enumerate(records):
        try:
            # 构造设备上传格式的数据 (去掉 base64 图片减小体积)
            payload = {
                "Mac_addr": r.get("Mac_addr", ""),
                "employee_number": r.get("employee_number", ""),
                "time": r.get("time", ""),
                "name": r.get("name", ""),
                "devicename": r.get("devicename", ""),
                "location": r.get("location", ""),
                "inout": r.get("inout", 0),
                "IdentifyType": r.get("IdentifyType", "0"),
                "resultStatus": r.get("resultStatus", 1),
                "id": r.get("id", ""),
                "SN": r.get("SN", ""),
                "icNum": r.get("icNum", ""),
                "sex": r.get("sex", ""),
                # 不传 face_base64 和 templatePhoto 以避免过大
                "_import": True,  # 标记为导入记录
            }
            
            data = json.dumps(payload).encode("utf-8")
            req = urllib.request.Request(
                gateway_url, data=data,
                headers={"Content-Type": "application/json"},
            )
            with urllib.request.urlopen(req, timeout=5) as resp:
                resp.read()
            success += 1
            
            if (i + 1) % 50 == 0:
                print(f"  进度: {i + 1}/{len(records)} (成功: {success})")
        except Exception as e:
            fail += 1
            if fail <= 3:
                print(f"  写入失败 #{fail}: {e}")
    
    print(f"\n✅ 大屏数据库写入完成: 成功 {success}, 失败 {fail}")

if __name__ == "__main__":
    main()
