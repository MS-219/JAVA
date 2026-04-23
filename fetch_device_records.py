#!/usr/bin/env python3
"""
从6台人脸识别设备上拉取 4月17日 到今天的识别记录，保存为 Excel。

步骤:
1. 对每台设备调用 setDeviceRecordRevert 触发重提记录
2. 设备会把记录推送到网关 (10.10.245.252:5389)
3. 等待推送完成后，通过网关 API 分页拉取所有数据
4. 保存为 Excel 文件
"""

import json
import time
import urllib.request
import urllib.error
import sys
import os
from datetime import datetime

# ========== 配置 ==========
GATEWAY_URL = "http://10.10.245.252:5389"
DEVICE_PASSWORD = "123456"
DEVICE_PORT = 8091

# 6台设备 (IP, MAC后缀, 名称)
DEVICES = [
    ("10.20.250.45", "61:B2", "大门东-中间 出口"),
    ("10.20.250.43", "61:CD", "大门东-东侧 入口"),
    ("10.20.250.44", "61:A5", "大门东-西侧 入口"),
    ("10.20.250.40", "61:D2", "大门西-东侧 入口"),
    ("10.20.250.37", "61:CF", "大门西-西侧 出口"),
    ("10.20.250.42", "61:D5", "大门西-中间 入口"),
]

# 时间范围
START_TIME = "2026-04-17 00:00:00"
END_TIME = "2026-04-23 23:59:59"

START_DATE = "2026-04-17"
END_DATE = "2026-04-23"


def post_json(url, data, timeout=15):
    """发送 POST JSON 请求"""
    payload = json.dumps(data).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=payload,
        headers={"Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        return {"error": str(e)}


def get_json(url, timeout=30):
    """发送 GET 请求"""
    try:
        with urllib.request.urlopen(url, timeout=timeout) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        return {"error": str(e)}


def trigger_record_revert():
    """对每台设备调用 setDeviceRecordRevert 触发记录重提"""
    print("=" * 60)
    print("步骤 1: 触发设备记录重提")
    print(f"时间范围: {START_TIME} ~ {END_TIME}")
    print("=" * 60)

    results = []
    for ip, mac, name in DEVICES:
        url = f"http://{ip}:{DEVICE_PORT}/setDeviceRecordRevert"
        data = {
            "password": DEVICE_PASSWORD,
            "data": {
                "startTime": START_TIME,
                "endTime": END_TIME,
            },
        }

        print(f"\n请求设备: {name} ({ip}) ...", end=" ")
        resp = post_json(url, data)

        if resp.get("result") == 0:
            print("✅ 成功")
        else:
            print(f"❌ 失败: {resp}")

        results.append((name, ip, resp))
        time.sleep(0.5)  # 避免同时请求太快

    return results


def wait_for_records(target_count=None, max_wait=180):
    """等待记录推送完成"""
    print("\n" + "=" * 60)
    print("步骤 2: 等待设备推送记录...")
    print("=" * 60)

    initial = get_json(f"{GATEWAY_URL}/dashboard/stats/paged?startDate={START_DATE}&endDate={END_DATE}&page=0&size=1")
    initial_count = initial.get("totalElements", 0)
    print(f"当前数据库中已有记录数: {initial_count}")

    last_count = initial_count
    stable_rounds = 0

    for i in range(max_wait // 5):
        time.sleep(5)
        resp = get_json(f"{GATEWAY_URL}/dashboard/stats/paged?startDate={START_DATE}&endDate={END_DATE}&page=0&size=1")
        current = resp.get("totalElements", 0)
        new_records = current - initial_count

        print(f"  [{(i + 1) * 5}s] 当前记录数: {current} (新增: {new_records})")

        if current == last_count:
            stable_rounds += 1
            if stable_rounds >= 6:  # 连续30秒无新记录
                print(f"\n已连续30秒无新记录，推送可能完成。")
                break
        else:
            stable_rounds = 0

        last_count = current

    final_count = last_count
    print(f"\n总共收到记录: {final_count} (新增: {final_count - initial_count})")
    return final_count


def export_to_excel(total_records):
    """从网关 API 拉取所有记录并保存为 Excel"""
    print("\n" + "=" * 60)
    print("步骤 3: 拉取数据并保存为 Excel")
    print("=" * 60)

    all_records = []
    page_size = 100
    total_pages = (total_records + page_size - 1) // page_size

    for page in range(total_pages):
        url = (
            f"{GATEWAY_URL}/dashboard/stats/paged"
            f"?startDate={START_DATE}&endDate={END_DATE}"
            f"&page={page}&size={page_size}"
        )
        resp = get_json(url, timeout=60)

        if "error" in resp:
            print(f"  页 {page + 1} 失败: {resp['error']}")
            continue

        records = resp.get("recentRecords", [])
        all_records.extend(records)
        print(f"  已拉取 第 {page + 1}/{total_pages} 页, 本页 {len(records)} 条, 累计 {len(all_records)} 条")

    if not all_records:
        print("没有获取到任何记录!")
        return

    # 保存为 Excel (使用 csv 格式, Excel 可直接打开)
    # 如果有 openpyxl 则用真正的 xlsx 格式
    try:
        import openpyxl
        save_as_xlsx(all_records)
    except ImportError:
        print("未安装 openpyxl, 尝试安装...")
        os.system(f"{sys.executable} -m pip install openpyxl -q")
        try:
            import openpyxl
            save_as_xlsx(all_records)
        except ImportError:
            print("openpyxl 安装失败, 改用 CSV 格式保存")
            save_as_csv(all_records)


def save_as_xlsx(records):
    """保存为真正的 xlsx Excel 文件"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "识别记录"

    # 表头
    headers = ["序号", "部门", "姓名", "识别时间", "设备位置", "进出方向", "人员类型"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # 类型映射
    type_map = {"teacher": "教职工", "student": "学生", "member": "会员", "guest": "访客"}

    # 数据行
    for idx, r in enumerate(records, 1):
        device_name = r.get("deviceName", "")
        direction = ""
        location = device_name

        if device_name and ("入口" in device_name or "出口" in device_name):
            parts = device_name.rsplit(" ", 1)
            if len(parts) == 2:
                location = parts[0]
                direction = parts[1]

        row_data = [
            idx,
            r.get("department", ""),
            r.get("name", ""),
            r.get("time", ""),
            location,
            direction,
            type_map.get(r.get("userType", ""), "访客"),
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 1, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if col in [1, 5, 6, 7] else "left")

    # 调整列宽
    col_widths = [8, 20, 15, 22, 22, 10, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 保存
    filename = f"识别记录_{START_DATE}_至_{END_DATE}.xlsx"
    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    wb.save(filepath)
    print(f"\n✅ 已保存为 Excel: {filepath}")
    print(f"   共 {len(records)} 条记录")


def save_as_csv(records):
    """备用: 保存为 CSV"""
    filename = f"识别记录_{START_DATE}_至_{END_DATE}.csv"
    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)

    with open(filepath, "w", encoding="utf-8-sig") as f:
        f.write("序号,部门,姓名,识别时间,设备位置,进出方向,人员类型\n")

        type_map = {"teacher": "教职工", "student": "学生", "member": "会员", "guest": "访客"}

        for idx, r in enumerate(records, 1):
            device_name = r.get("deviceName", "")
            direction = ""
            location = device_name
            if device_name and ("入口" in device_name or "出口" in device_name):
                parts = device_name.rsplit(" ", 1)
                if len(parts) == 2:
                    location = parts[0]
                    direction = parts[1]

            row = [
                str(idx),
                r.get("department", ""),
                r.get("name", ""),
                r.get("time", ""),
                location,
                direction,
                type_map.get(r.get("userType", ""), "访客"),
            ]
            f.write(",".join(f'"{v}"' for v in row) + "\n")

    print(f"\n✅ 已保存为 CSV: {filepath}")
    print(f"   共 {len(records)} 条记录")


def main():
    print(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"目标: 拉取 {START_DATE} 至 {END_DATE} 的识别记录\n")

    # 先检查当前已有多少记录
    resp = get_json(f"{GATEWAY_URL}/dashboard/stats/paged?startDate={START_DATE}&endDate={END_DATE}&page=0&size=1")
    current_count = resp.get("totalElements", 0)
    print(f"数据库中当前已有 {current_count} 条记录 (日期范围内)")

    # 步骤1: 触发设备记录重提
    trigger_record_revert()

    # 步骤2: 等待记录推送完成
    total = wait_for_records(max_wait=180)

    # 步骤3: 导出为 Excel
    export_to_excel(total)

    print(f"\n完成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    main()
